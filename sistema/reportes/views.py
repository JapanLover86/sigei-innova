from decimal import Decimal
from io import BytesIO

from django.http import HttpResponse
from django.shortcuts import render
from django.template.loader import render_to_string
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from xhtml2pdf import pisa

from inventario.models import MovimientoStock
from productos.models import Producto
from usuarios.decorators import rol_requerido


def generar_pdf(template_name, contexto, nombre_archivo):
    html = render_to_string(template_name, contexto)

    resultado = BytesIO()

    pdf = pisa.CreatePDF(
        html,
        dest=resultado,
        encoding="UTF-8",
    )

    if pdf.err:
        return HttpResponse(
            "Ocurrió un error al generar el PDF.",
            status=500,
        )

    respuesta = HttpResponse(
        resultado.getvalue(),
        content_type="application/pdf",
    )

    respuesta[
        "Content-Disposition"
    ] = f'attachment; filename="{nombre_archivo}"'

    return respuesta


def generar_excel(nombre_archivo, titulo, encabezados, filas):
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Reporte"

    hoja.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=len(encabezados),
    )

    celda_titulo = hoja.cell(row=1, column=1)
    celda_titulo.value = titulo
    celda_titulo.font = Font(bold=True, size=14)
    celda_titulo.alignment = Alignment(horizontal="center")

    for columna, encabezado in enumerate(encabezados, start=1):
        celda = hoja.cell(row=3, column=columna)
        celda.value = encabezado
        celda.font = Font(bold=True)
        celda.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )
        celda.alignment = Alignment(horizontal="center")

    for fila_numero, fila in enumerate(filas, start=4):
        for columna, valor in enumerate(fila, start=1):
            celda = hoja.cell(
                row=fila_numero,
                column=columna,
            )
            celda.value = valor
            celda.alignment = Alignment(vertical="center")

    # Ajustar ancho de columnas sin recorrer las celdas combinadas del título.
    for indice_columna in range(1, len(encabezados) + 1):
        letra_columna = get_column_letter(indice_columna)
        ancho = 15

        for fila_numero in range(3, hoja.max_row + 1):
            valor = hoja.cell(
                row=fila_numero,
                column=indice_columna,
            ).value

            if valor is not None:
                ancho = max(ancho, len(str(valor)) + 2)

        hoja.column_dimensions[letra_columna].width = min(ancho, 40)

    hoja.freeze_panes = "A4"

    resultado = BytesIO()
    libro.save(resultado)
    resultado.seek(0)

    respuesta = HttpResponse(
        resultado.getvalue(),
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
    )

    respuesta[
        "Content-Disposition"
    ] = f'attachment; filename="{nombre_archivo}"'

    return respuesta


@rol_requerido("ADMIN", "GERENTE", "ALMACENERO")
def reporte_inicio(request):
    return render(
        request,
        "reportes/inicio.html",
    )


@rol_requerido("ADMIN", "GERENTE", "ALMACENERO")
def inventario_pdf(request):
    productos = Producto.objects.filter(
        estado=True,
    ).select_related(
        "categoria",
        "marca",
    ).order_by(
        "nombre",
    )

    valor_total = sum(
        (
            (producto.stock_actual or Decimal("0.00"))
            * (producto.precio_compra or Decimal("0.00"))
        )
        for producto in productos
    )

    contexto = {
        "productos": productos,
        "valor_total": valor_total,
        "fecha_reporte": timezone.localtime(),
    }

    return generar_pdf(
        "reportes/inventario_pdf.html",
        contexto,
        "reporte_inventario_sigei.pdf",
    )


@rol_requerido("ADMIN", "GERENTE", "ALMACENERO")
def inventario_excel(request):
    productos = Producto.objects.filter(
        estado=True,
    ).select_related(
        "categoria",
        "marca",
    ).order_by(
        "nombre",
    )

    encabezados = [
        "Código",
        "Producto",
        "Categoría",
        "Marca",
        "Unidad",
        "Stock actual",
        "Stock mínimo",
        "Precio compra",
        "Precio venta",
        "Estado de stock",
    ]

    filas = []

    for producto in productos:
        if producto.stock_actual <= 0:
            estado_stock = "Sin stock"
        elif producto.stock_actual <= producto.stock_minimo:
            estado_stock = "Stock bajo"
        else:
            estado_stock = "Disponible"

        filas.append(
            [
                producto.codigo,
                producto.nombre,
                producto.categoria.nombre,
                producto.marca.nombre,
                producto.unidad,
                float(producto.stock_actual),
                float(producto.stock_minimo),
                float(producto.precio_compra),
                float(producto.precio_venta),
                estado_stock,
            ]
        )

    return generar_excel(
        "reporte_inventario_sigei.xlsx",
        "Reporte de Inventario - SIGEI Innova",
        encabezados,
        filas,
    )


@rol_requerido("ADMIN", "GERENTE", "ALMACENERO")
def kardex_pdf(request):
    movimientos = MovimientoStock.objects.select_related(
        "producto",
    ).order_by(
        "-fecha",
        "-id_movimiento",
    )

    contexto = {
        "movimientos": movimientos,
        "fecha_reporte": timezone.localtime(),
    }

    return generar_pdf(
        "reportes/kardex_pdf.html",
        contexto,
        "reporte_kardex_sigei.pdf",
    )


@rol_requerido("ADMIN", "GERENTE", "ALMACENERO")
def kardex_excel(request):
    movimientos = MovimientoStock.objects.select_related(
        "producto",
    ).order_by(
        "-fecha",
        "-id_movimiento",
    )

    encabezados = [
        "Fecha",
        "Código",
        "Producto",
        "Tipo",
        "Referencia",
        "Cantidad",
        "Stock anterior",
        "Stock actual",
        "Observación",
    ]

    filas = []

    for movimiento in movimientos:
        filas.append(
            [
                movimiento.fecha.strftime("%d/%m/%Y %H:%M"),
                movimiento.producto.codigo,
                movimiento.producto.nombre,
                movimiento.tipo_movimiento,
                movimiento.referencia or "-",
                float(movimiento.cantidad),
                float(movimiento.stock_anterior),
                float(movimiento.stock_actual),
                movimiento.observacion or "-",
            ]
        )

    return generar_excel(
        "reporte_kardex_sigei.xlsx",
        "Reporte de Kardex - SIGEI Innova",
        encabezados,
        filas,
    )